# This script is for judging programming homework.
# Usage: Put this script in the same dir as the homework and then run this script.

import subprocess
import os
import time
import zipfile
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors

WA = 0
AC = 1
TLE = 2
RTE = 3


def check_once(executable, input, output):
    '''check the executable for one sample

    :param executable: String. the path of the executable
    :param input: String. sample input
    :param output: String. sample output
    :return: The constants, such as AC, TLE, etc.
    '''
    try:
        ret = subprocess.run(
            executable,
            input=input,
            stdout=subprocess.PIPE,
            timeout=1,
            encoding='gbk')
        ret.check_returncode()
    except subprocess.TimeoutExpired:
        return TLE
    except subprocess.CalledProcessError:
        return RTE

    if (ret.stdout.replace('\r\n', '\n').strip('\n') != output
            and ret.stdout.replace('\r\n', '\n').strip('\n') + '\n' != output):  # 忽略末尾换行
        return WA
    else:
        return AC


def check(executable, inputs, outputs):
    '''check the executable for multiple samples

    :param executable: String. the path of the executable
    :param inputs: list of String. sample input
    :param outputs: list of String. sample output
    :return: list of the constants
    '''
    return [check_once(executable, inputs[i], outputs[i]) for i in range(len(inputs))]


def get_samples(n):
    '''get samples for judging an executable

    :param n: Integer. exercise number
    :return: Tuple. (list of inputs, list of outputs)
    '''
    with open('input_' + str(n) + '.txt', "r") as inputs_file:
        inputs = []

        row_num = inputs_file.readline()
        while row_num != '' and row_num != None:
            input = ""
            for i in range(int(row_num)):
                input += inputs_file.readline()
            inputs.append(input)

            row_num = inputs_file.readline()

    with open('output_' + str(n) + '.txt', "r") as outputs_file:
        outputs = []

        row_num = outputs_file.readline()
        while row_num != '' and row_num != None:
            output = ""
            for i in range(int(row_num)):
                output += outputs_file.readline()
            outputs.append(output)

            row_num = outputs_file.readline()

    return inputs, outputs

def extract_zip(file):
    '''extract a zip file to current directory.

    :param file: String, zip file name
    :return: no return
    '''
    with zipfile.ZipFile(file, 'r') as zf:
        for name in zf.namelist():
            if not os.path.exists(name.encode('cp437').decode('gbk')):
                os.renames(zf.extract(name), name.encode('cp437').decode('gbk'))


def main_zips():
    '''main program for zip files structure

    :return: no return
    '''
    files = os.listdir('.')
    zipfilenames = []
    dirnames = []

    # ----------Extract the zip files----------
    for file in files:
        if not os.path.isdir(file):
            if zipfile.is_zipfile(file) and file.split('.')[1] == 'zip':
                extract_zip(file)
                dirnames.append(file.split('.')[0])
                zipfilenames.append(file)

    # ----------check for exercises----------
    main_folders()

    # ----------delete the extracted files and move the judged one to "Judged" dir----------
    for dirname in dirnames:
        shutil.rmtree(dirname)
    if not os.path.exists('Judged'):
        os.mkdir('Judged')
    for zipfilename in zipfilenames:
        shutil.move(zipfilename, 'Judged')


def main_folders():
    '''main program for folder structure

    :return: no return
    '''
    WORKBOOK_PATH = '新生c.xlsx'
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb['result']
    cell_range = ws['A3':'A77']
    names = [cell[0] for cell in cell_range]

    files = os.listdir('.')
    for file in files:
        if os.path.isdir(file):
            # ----------check for exercises----------
            dirname = file
            participant_name = dirname.split(' ')[0].split('+')[0]

            for i in range(1, 5):
                inputs, outputs = get_samples(i)
                ws2 = wb['2_' + str(i)]
                print('Judging ' + participant_name + ' for exercise ' + str(i))

                begin_time = time.perf_counter()
                result = check('./' + dirname + '/' + str(i) + '/' + str(i) + '.exe', inputs, outputs)
                end_time = time.perf_counter()

                accuracy = result.count(AC) / len(result)
                time_consumption = end_time - begin_time
                print('Successfully judged ' + participant_name + ' for exercise ' + str(i))
                print('accuracy\ttime')
                print('{:.3f}'.format(accuracy) + '\t' + '{:.3f}'.format(time_consumption))

                for name in names:
                    if name.value == participant_name:
                        ws.cell(row=name.row, column=name.col_idx + 3 * (i - 1) + 1).value = accuracy
                        ws.cell(row=name.row, column=name.col_idx + 3 * (i - 1) + 2).value = time_consumption
                        for i in range(len(result)):
                            if result[i] == AC:
                                ws2.cell(row=name.row, column=name.col_idx + i + 1).value = 'AC'
                                ws2.cell(row=name.row, column=name.col_idx + i + 1).font = Font(color=colors.GREEN)
                            elif result[i] == WA:
                                ws2.cell(row=name.row, column=name.col_idx + i + 1).value = 'WA'
                                ws2.cell(row=name.row, column=name.col_idx + i + 1).font = Font(color=colors.RED)
                            elif result[i] == TLE:
                                ws2.cell(row=name.row, column=name.col_idx + i + 1).value = 'TLE'
                                ws2.cell(row=name.row, column=name.col_idx + i + 1).font = Font(color=colors.DARKYELLOW)
                            elif result[i] == RTE:
                                ws2.cell(row=name.row, column=name.col_idx + i + 1).value = 'RTE'
                                ws2.cell(row=name.row, column=name.col_idx + i + 1).font = Font(color=colors.DARKYELLOW)
                        wb.save(WORKBOOK_PATH)
                        print('Successfully saved ' + participant_name + ' for exercise ' + str(i) + '\n')
                        break

    wb.close()


if __name__ == '__main__':
    main_folders()